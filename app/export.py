"""
Export module for Portfolio Optimizer - Native Excel Charts Edition

This module handles all data export functionality (CSV, Excel) using native Excel
charts created by openpyxl, eliminating the Kaleido/Chrome dependency that causes
failures on Streamlit Cloud.

Charts are actual editable Excel objects, not embedded images, which provides better
compatibility and professional appearance.
"""

from typing import Dict
from io import BytesIO
import pandas as pd

from openpyxl import Workbook  # type: ignore
from openpyxl.chart import PieChart, LineChart, Reference as ChartReference  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore

from app.config import EXCEL_MAX_COLUMN_WIDTH  # type: ignore


def generate_excel_multiple_portfolios(optimized_data: dict) -> BytesIO:
    """
    Generate Excel workbook with one worksheet per portfolio.
    
    Each portfolio gets its own worksheet containing:
    - Metrics comparison table (Portfolio vs Benchmark)
    - Holdings table
    - Portfolio allocation pie chart
    - Cumulative returns line chart
    
    Args:
        optimized_data: Dictionary from prepare_multiple_portfolio_data containing:
            - portfolios: Dict mapping portfolio_type -> portfolio_data
            - benchmark: Benchmark data (cum_rets, daily_rets, perf, etc.)
            - prices: Price history DataFrame
            - period_days: Number of days in analysis period
            - chart_data: DataFrame with cumulative returns for all portfolios
    
    Returns:
        BytesIO object containing the Excel workbook with multiple worksheets
    
    Raises:
        ValueError: If workbook generation fails
    """
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Extract data
        portfolios = optimized_data.get("portfolios", {})
        period_days = optimized_data.get("period_days", 0)
        chart_data = optimized_data.get("chart_data", pd.DataFrame())
        
        # Portfolio type to display name mapping
        portfolio_display_names = {
            "max_sharpe": "Max Sharpe",
            "min_variance": "Min Volatility",
            "efficient_return": "Efficient Return",
            "efficient_risk": "Efficient Risk",
            "efficient_te": "Efficient TE",
        }
        
        # Create worksheet for each portfolio
        for portfolio_type, portfolio_data in portfolios.items():
            # Get display name for worksheet
            display_name = portfolio_display_names.get(portfolio_type, portfolio_type.replace("_", " ").title())
            
            # Create worksheet (truncate to 31 chars for Excel sheet name limit)
            ws_name = display_name[:31]
            ws = wb.create_sheet(title=ws_name)
            
            # =====================================================================
            # SECTION 1: TITLE AND METADATA
            # =====================================================================
            
            ws['A1'] = f'{display_name} Portfolio Analysis'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Analysis Period: {period_days} days'
            
            # =====================================================================
            # SECTION 2: METRICS COMPARISON TABLE
            # =====================================================================
            
            startrow_metrics = 3
            
            comparison_df = portfolio_data.get("comparison_df", pd.DataFrame())
            
            # Write metric names column
            ws.cell(row=startrow_metrics, column=1, value="Metric").font = Font(bold=True)
            for col_name in comparison_df.columns:
                col_idx = list(comparison_df.columns).index(col_name) + 2
                ws.cell(row=startrow_metrics, column=col_idx, value=col_name).font = Font(bold=True)
            
            # Write metric rows
            for row_offset, (metric_name, row_data) in enumerate(comparison_df.iterrows(), 1):
                ws.cell(row=startrow_metrics + row_offset, column=1, value=metric_name)
                for col_idx, col_name in enumerate(comparison_df.columns, 2):
                    value = row_data[col_name]
                    ws.cell(row=startrow_metrics + row_offset, column=col_idx, value=value)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            for col_idx in range(2, 2 + len(comparison_df.columns)):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18
            
            metrics_last_row = startrow_metrics + len(comparison_df)
            
            # =====================================================================
            # SECTION 3: HOLDINGS TABLE
            # =====================================================================
            
            holdings_display = portfolio_data.get("holdings_df", pd.DataFrame())
            holdings_header_row = metrics_last_row + 3
            holdings_data_row = holdings_header_row + 1
            
            ws.cell(row=holdings_header_row, column=1, value="Holdings").font = Font(bold=True, size=12)
            
            # Write holdings table header
            for col_idx, col_name in enumerate(holdings_display.columns, 1):
                cell = ws.cell(row=holdings_data_row, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Write holdings table data
            for row_idx, row_data in enumerate(holdings_display.values, holdings_data_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Adjust column widths
            for col_idx, col_name in enumerate(holdings_display.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(col_name))
                for val in holdings_display[col_name]:
                    max_length = max(max_length, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, EXCEL_MAX_COLUMN_WIDTH)
            
            holdings_last_row = holdings_data_row + len(holdings_display)
            
            # =====================================================================
            # SECTION 4: PIE CHART DATA AND CHART
            # =====================================================================
            
            pie_chart_section_row = holdings_last_row + 3
            pie_data_row = pie_chart_section_row + 1
            
            ws.cell(row=pie_chart_section_row, column=1, value="Portfolio Allocation").font = Font(bold=True, size=12)
            
            # Prepare data for pie chart (filter zero weights)
            pie_data_col_A = 1  # Ticker
            pie_data_col_B = 2  # Weight
            
            ws.cell(row=pie_data_row, column=pie_data_col_A, value="Ticker").font = Font(bold=True)
            ws.cell(row=pie_data_row, column=pie_data_col_B, value="Weight").font = Font(bold=True)
            
            # Extract non-zero weights
            weights = portfolio_data.get("weights", {})
            pie_data_current_row = pie_data_row + 1
            pie_data_max_row = pie_data_current_row
            
            for ticker, weight in sorted(weights.items(), key=lambda x: x[1], reverse=True):
                if weight > 0.001:  # At least 0.1%
                    ws.cell(row=pie_data_current_row, column=pie_data_col_A, value=ticker)
                    ws.cell(row=pie_data_current_row, column=pie_data_col_B, value=weight)
                    pie_data_max_row = pie_data_current_row
                    pie_data_current_row += 1
            
            # Create pie chart
            if pie_data_max_row > pie_data_row:
                pie_chart = PieChart()
                pie_chart.title = "Portfolio Weights"
                pie_chart.style = 10
                pie_chart.height = 10
                pie_chart.width = 15
                
                # Data references for pie chart
                labels_ref = ChartReference(
                    ws,
                    min_col=pie_data_col_A,
                    min_row=pie_data_row + 1,
                    max_row=pie_data_max_row
                )
                data_ref = ChartReference(
                    ws,
                    min_col=pie_data_col_B,
                    min_row=pie_data_row,
                    max_row=pie_data_max_row
                )
                
                pie_chart.add_data(data_ref, titles_from_data=True)
                pie_chart.set_categories(labels_ref)
                
                ws.add_chart(pie_chart, f'D{pie_data_row}')
            
            # =====================================================================
            # SECTION 5: LINE CHART DATA AND CHART
            # =====================================================================
            
            line_chart_section_row = pie_data_current_row + 2
            line_data_row = line_chart_section_row + 1
            
            # Write cumulative returns data to worksheet for line chart
            line_data_col_date = 3   # Column C
            line_data_col_pf = 4     # Column D
            line_data_col_bmk = 5    # Column E
            
            ws.cell(row=line_data_row, column=line_data_col_date, value="Date").font = Font(bold=True)
            ws.cell(row=line_data_row, column=line_data_col_pf, value=display_name).font = Font(bold=True)
            ws.cell(row=line_data_row, column=line_data_col_bmk, value="Benchmark").font = Font(bold=True)
            
            # Write chart data - use chart_data which includes all portfolios
            line_data_current_row = line_data_row + 1
            line_data_max_row = line_data_current_row
            
            for date_label, row_data in chart_data.iterrows():
                ws.cell(row=line_data_current_row, column=line_data_col_date, value=date_label)
                ws.cell(row=line_data_current_row, column=line_data_col_pf, value=row_data.get(display_name, 0))
                ws.cell(row=line_data_current_row, column=line_data_col_bmk, value=row_data.get('Benchmark', 0))
                line_data_max_row = line_data_current_row
                line_data_current_row += 1
            
            # Create line chart
            if line_data_max_row > line_data_row:
                line_chart = LineChart()
                line_chart.title = "Cumulative Returns Comparison"
                line_chart.style = 12
                line_chart.y_axis.title = "Cumulative Return"
                line_chart.x_axis.title = "Date"
                line_chart.height = 10
                line_chart.width = 18
                line_chart.legend.position = 'r'
                
                # Data references for line chart
                x_data_ref = ChartReference(
                    ws,
                    min_col=line_data_col_date,
                    min_row=line_data_row + 1,
                    max_row=line_data_max_row
                )
                pf_data_ref = ChartReference(
                    ws,
                    min_col=line_data_col_pf,
                    min_row=line_data_row,
                    max_row=line_data_max_row
                )
                bmk_data_ref = ChartReference(
                    ws,
                    min_col=line_data_col_bmk,
                    min_row=line_data_row,
                    max_row=line_data_max_row
                )
                
                line_chart.add_data(pf_data_ref, titles_from_data=True)
                line_chart.add_data(bmk_data_ref, titles_from_data=True)
                line_chart.set_categories(x_data_ref)
                
                ws.add_chart(line_chart, f'C{line_chart_section_row + 1}')
        
        # =====================================================================
        # SAVE WORKBOOK TO BUFFER
        # =====================================================================
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        raise ValueError(f"Failed to generate Excel export with multiple portfolios: {str(e)}")


def generate_excel_full_page(
    comparison_df: pd.DataFrame,
    holdings_display: pd.DataFrame,
    period_days: int,
    chart_data: pd.DataFrame,
    weights: Dict[str, float]
) -> BytesIO:
    """
    Generate comprehensive Excel workbook with metrics, holdings, and native Excel charts.
    
    Uses openpyxl's built-in charting to create professional, editable charts without
    requiring Kaleido or Chrome. This solution works perfectly on Streamlit Cloud.
    
    Layout:
    - Row 1: Title
    - Row 2: Period info
    - Row 3+: Summary returns table
    - Below: Holdings table
    - Bottom: Pie chart (portfolio allocation) and Line chart (cumulative returns)
    
    Args:
        comparison_df: Performance comparison DataFrame (Portfolio vs Benchmark metrics)
        holdings_display: Holdings table DataFrame with Ticker, Security, GICS Sector, Weights
        period_days: Number of days in the analysis period
        chart_data: DataFrame with cumulative returns time series
                   Index: dates, Columns: ['Max Sharpe PF', 'Benchmark']
        weights: Dict mapping ticker -> weight for portfolio allocation pie chart
    
    Returns:
        BytesIO object containing the Excel workbook
    
    Raises:
        ValueError: If workbook generation fails
    """
    try:
        # Create workbook
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "MaxSharpe"
        
        # =====================================================================
        # SECTION 1: TITLE AND METADATA
        # =====================================================================
        
        ws_main['A1'] = 'Maximum Sharpe Portfolio Analysis'
        ws_main['A1'].font = Font(bold=True, size=14)
        ws_main['A2'] = f'Analysis Period: {period_days} days'
        
        # =====================================================================
        # SECTION 2: METRICS COMPARISON TABLE
        # =====================================================================
        
        startrow_metrics = 3
        
        # Write metric names column
        ws_main.cell(row=startrow_metrics, column=1, value="Metric").font = Font(bold=True)
        for col_name in comparison_df.columns:
            col_idx = list(comparison_df.columns).index(col_name) + 2
            ws_main.cell(row=startrow_metrics, column=col_idx, value=col_name).font = Font(bold=True)
        
        # Write metric rows
        for row_offset, (metric_name, row_data) in enumerate(comparison_df.iterrows(), 1):
            ws_main.cell(row=startrow_metrics + row_offset, column=1, value=metric_name)
            for col_idx, col_name in enumerate(comparison_df.columns, 2):
                value = row_data[col_name]
                ws_main.cell(row=startrow_metrics + row_offset, column=col_idx, value=value)
        
        # Adjust column widths
        ws_main.column_dimensions['A'].width = 25
        for col_idx in range(2, 2 + len(comparison_df.columns)):
            ws_main.column_dimensions[get_column_letter(col_idx)].width = 18
        
        metrics_last_row = startrow_metrics + len(comparison_df)
        
        # =====================================================================
        # SECTION 3: HOLDINGS TABLE
        # =====================================================================
        
        holdings_header_row = metrics_last_row + 3
        holdings_data_row = holdings_header_row + 1
        
        ws_main.cell(row=holdings_header_row, column=1, value="Holdings").font = Font(bold=True, size=12)
        
        # Write holdings table header
        for col_idx, col_name in enumerate(holdings_display.columns, 1):
            cell = ws_main.cell(row=holdings_data_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write holdings table data
        for row_idx, row_data in enumerate(holdings_display.values, holdings_data_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_main.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, col_name in enumerate(holdings_display.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(col_name))
            for val in holdings_display[col_name]:
                max_length = max(max_length, len(str(val)))
            ws_main.column_dimensions[col_letter].width = min(max_length + 2, EXCEL_MAX_COLUMN_WIDTH)
        
        holdings_last_row = holdings_data_row + len(holdings_display)
        
        # =====================================================================
        # SECTION 4: PIE CHART DATA AND CHART
        # =====================================================================
        
        pie_chart_section_row = holdings_last_row + 3
        pie_data_row = pie_chart_section_row + 1
        
        ws_main.cell(row=pie_chart_section_row, column=1, value="Portfolio Allocation & Performance").font = Font(bold=True, size=12)
        
        # Prepare data for pie chart (filter zero weights)
        pie_data_col_A = 1  # Ticker
        pie_data_col_B = 2  # Weight
        
        ws_main.cell(row=pie_data_row, column=pie_data_col_A, value="Ticker").font = Font(bold=True)
        ws_main.cell(row=pie_data_row, column=pie_data_col_B, value="Weight").font = Font(bold=True)
        
        # Extract non-zero weights
        pie_data_current_row = pie_data_row + 1
        pie_data_max_row = pie_data_current_row
        
        for ticker, weight in sorted(weights.items(), key=lambda x: x[1], reverse=True):
            if weight > 0.001:  # At least 0.1%
                ws_main.cell(row=pie_data_current_row, column=pie_data_col_A, value=ticker)
                ws_main.cell(row=pie_data_current_row, column=pie_data_col_B, value=weight)
                pie_data_max_row = pie_data_current_row
                pie_data_current_row += 1
        
        # Create pie chart
        if pie_data_max_row > pie_data_row:
            pie_chart = PieChart()
            pie_chart.title = "Portfolio Weights"
            pie_chart.style = 10
            pie_chart.height = 10
            pie_chart.width = 15
            
            # Data references for pie chart
            labels_ref = ChartReference(
                ws_main,
                min_col=pie_data_col_A,
                min_row=pie_data_row + 1,
                max_row=pie_data_max_row
            )
            data_ref = ChartReference(
                ws_main,
                min_col=pie_data_col_B,
                min_row=pie_data_row,
                max_row=pie_data_max_row
            )
            
            pie_chart.add_data(data_ref, titles_from_data=True)
            pie_chart.set_categories(labels_ref)
            
            ws_main.add_chart(pie_chart, f'D{pie_data_row}')
        
        # =====================================================================
        # SECTION 5: LINE CHART DATA AND CHART
        # =====================================================================
        
        line_chart_section_row = pie_data_current_row + 2
        line_data_row = line_chart_section_row + 1
        
        # Write cumulative returns data to worksheet for line chart
        # Column C = Date/Index, Column D = Portfolio, Column E = Benchmark
        line_data_col_date = 3   # Column C
        line_data_col_pf = 4     # Column D
        line_data_col_bmk = 5    # Column E
        
        ws_main.cell(row=line_data_row, column=line_data_col_date, value="Date").font = Font(bold=True)
        ws_main.cell(row=line_data_row, column=line_data_col_pf, value="Max Sharpe PF").font = Font(bold=True)
        ws_main.cell(row=line_data_row, column=line_data_col_bmk, value="Benchmark").font = Font(bold=True)
        
        # Write chart data
        line_data_current_row = line_data_row + 1
        line_data_max_row = line_data_current_row
        
        for date_label, row_data in chart_data.iterrows():
            ws_main.cell(row=line_data_current_row, column=line_data_col_date, value=date_label)
            ws_main.cell(row=line_data_current_row, column=line_data_col_pf, value=row_data.get('Max Sharpe PF', 0))
            ws_main.cell(row=line_data_current_row, column=line_data_col_bmk, value=row_data.get('Benchmark', 0))
            line_data_max_row = line_data_current_row
            line_data_current_row += 1
        
        # Create line chart
        if line_data_max_row > line_data_row:
            line_chart = LineChart()
            line_chart.title = "Cumulative Returns Comparison"
            line_chart.style = 12
            line_chart.y_axis.title = "Cumulative Return"
            line_chart.x_axis.title = "Date"
            line_chart.height = 10
            line_chart.width = 18
            line_chart.legend.position = 'r'
            
            # Data references for line chart
            x_data_ref = ChartReference(
                ws_main,
                min_col=line_data_col_date,
                min_row=line_data_row + 1,
                max_row=line_data_max_row
            )
            pf_data_ref = ChartReference(
                ws_main,
                min_col=line_data_col_pf,
                min_row=line_data_row,
                max_row=line_data_max_row
            )
            bmk_data_ref = ChartReference(
                ws_main,
                min_col=line_data_col_bmk,
                min_row=line_data_row,
                max_row=line_data_max_row
            )
            
            line_chart.add_data(pf_data_ref, titles_from_data=True)
            line_chart.add_data(bmk_data_ref, titles_from_data=True)
            line_chart.set_categories(x_data_ref)
            
            ws_main.add_chart(line_chart, f'C{line_chart_section_row + 1}')
        

        
        # =====================================================================
        # SAVE WORKBOOK TO BUFFER
        # =====================================================================
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        raise ValueError(f"Failed to generate Excel export: {str(e)}")
